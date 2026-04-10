import openpyxl
import uuid
from icalendar import Calendar, Event, Alarm
from datetime import datetime, timedelta
import pytz

def create_ics_from_excel(excel_file, output_file, calendar_name):
    # Load file Excel
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb.active
    
    cal = Calendar()
    # Các thông tin header của Calendar
    cal.add('prodid', '-//iCalendar//mxm//VN')
    cal.add('name', f'{calendar_name}')
    cal.add('version', '2.0')
    cal.add('calscale', 'GREGORIAN')
    cal.add('method', 'PUBLISH')
    
    # Múi giờ Việt Nam
    tz_vn = pytz.timezone('Asia/Ho_Chi_Minh')
    current_year = 2026

    # Duyệt dữ liệu theo bước nhảy 4 dòng
    for i in range(1, sheet.max_row + 1, 4):
        time_val = sheet.cell(row=i, column=1).value
        t1 = sheet.cell(row=i+1, column=1).value
        t2 = sheet.cell(row=i+2, column=1).value

        if not time_val or "." not in str(time_val):
            continue

        try:
            # 1. Xử lý thời gian và Timezone
            time_str = str(time_val).strip()
            dt_obj = datetime.strptime(f"{time_str}.{current_year}", "%d.%m. %H:%M.%Y")
            dt_localized = tz_vn.localize(dt_obj)
            summary_text = f"{t1} - {t2}"

            # 3. Tạo Event theo format yêu cầu
            event = Event()
            event.add('uid', str(uuid.uuid4()).upper()) # Tạo UID ngẫu nhiên
            event.add('dtstamp', datetime.now(pytz.utc))
            event.add('dtstart', dt_localized)
            event.add('dtend', dt_localized + timedelta(hours=2))
            event.add('summary', summary_text)
            event.add('created', datetime.now(pytz.utc))
            event.add('description', f'Next match: {summary_text}')
            event.categories = ["eSport"]

            # --- Alarm 1: Trước 1 ngày (-P1D) ---
            # alarm_1d = Alarm()
            # alarm_1d.add('uid', str(uuid.uuid4()).upper())
            # alarm_1d.add('action', 'DISPLAY')
            # alarm_1d.add('description', f"Nhắc nhở: {summary_text} sẽ diễn ra vào ngày mai!")
            # alarm_1d.add('trigger', timedelta(days=-1))
            # event.add_component(alarm_1d)

            # --- Alarm 2: Trước 10 phút (-PT10M) ---
            alarm_10m = Alarm()
            alarm_10m.add('uid', str(uuid.uuid4()).upper())
            alarm_10m.add('action', 'DISPLAY')
            alarm_10m.add('description', f"Sắp bắt đầu: {summary_text}")
            alarm_10m.add('trigger', timedelta(minutes=-10))
            event.add_component(alarm_10m)

            cal.add_component(event)

        except Exception as e:
            print(f"Bỏ qua dòng {i} do lỗi: {e}")

    # Xuất file dạng binary ical
    with open(output_file, 'wb') as f:
        f.write(cal.to_ical())
    
    print(f"Xong! Đã tạo file: {output_file}")

if __name__ == "__main__":
    calendar_name = "LCK 2026"
    excel_file = "C:\\Python\\excel2ics\\lck-schedule.xlsx"
    ics_file = 'lck-schedule-2026.ics'
    create_ics_from_excel(excel_file, ics_file, calendar_name)